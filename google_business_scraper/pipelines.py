import openpyxl
from openpyxl import Workbook
import os
from itemadapter import ItemAdapter

class ValidationPipeline:
    def process_item(self, item, spider):
        adapter = ItemAdapter(item)

        # Validar se os campos obrigatórios estão presentes
        if not adapter.get('name'):
            spider.logger.warning(f"Item sem nome descartado: {item}")
            return None

        return item

class ExcelExportPipeline:
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.row_count = 1
        self.filename = None

    def open_spider(self, spider):
        # Criar diretório data se não existir
        os.makedirs('data', exist_ok=True)

        # Determinar nome do arquivo baseado na busca
        search_query = getattr(spider, 'search_query', 'salao_de_beleza_atibaia')

        # Limpar o nome do arquivo
        safe_filename = search_query.replace(' ', '_').replace(',', '').replace('/', '_').lower()
        safe_filename = ''.join(c for c in safe_filename if c.isalnum() or c in ('_', '-'))

        # Usar configuração do settings se disponível
        settings_filename = spider.settings.get('EXCEL_FILENAME')
        if settings_filename:
            self.filename = f"data/{settings_filename}"
        else:
            self.filename = f"data/{safe_filename}.xlsx"

        # Criar workbook
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Resultados da Busca"

        # Adicionar cabeçalhos
        headers = ['Nome', 'Avaliação', 'Número de Avaliações', 'Localização', 'Categoria', 'URL']
        for col, header in enumerate(headers, 1):
            self.worksheet.cell(row=1, column=col, value=header)

        spider.logger.info(f"Pipeline Excel iniciado - arquivo será salvo em {self.filename}")

    def close_spider(self, spider):
        if self.workbook and self.filename:
            try:
                self.workbook.save(self.filename)
                spider.logger.info(f"Arquivo Excel salvo com sucesso: {self.filename}")
                spider.logger.info(f"Total de itens salvos: {self.row_count - 1}")

                # Verificar se o arquivo foi criado
                if os.path.exists(self.filename):
                    file_size = os.path.getsize(self.filename)
                    spider.logger.info(f"Arquivo confirmado - Tamanho: {file_size} bytes")
                else:
                    spider.logger.error(f"Arquivo não encontrado após salvamento: {self.filename}")

            except Exception as e:
                spider.logger.error(f"Erro ao salvar arquivo Excel: {e}")
                # Tentar salvar no diretório raiz como fallback
                try:
                    fallback_filename = os.path.basename(self.filename)
                    self.workbook.save(fallback_filename)
                    spider.logger.info(f"Arquivo salvo no diretório raiz: {fallback_filename}")
                except Exception as e2:
                    spider.logger.error(f"Erro ao salvar no diretório raiz: {e2}")

    def process_item(self, item, spider):
        if item:
            adapter = ItemAdapter(item)
            self.row_count += 1

            # Adicionar dados à planilha
            self.worksheet.cell(row=self.row_count, column=1, value=adapter.get('name', ''))
            self.worksheet.cell(row=self.row_count, column=2, value=adapter.get('rating', ''))
            self.worksheet.cell(row=self.row_count, column=3, value=adapter.get('review_count', ''))
            self.worksheet.cell(row=self.row_count, column=4, value=adapter.get('location', ''))
            self.worksheet.cell(row=self.row_count, column=5, value=adapter.get('category', ''))
            self.worksheet.cell(row=self.row_count, column=6, value=adapter.get('url', ''))

            spider.logger.info(f"Item adicionado à planilha: {adapter.get('name', 'Nome não encontrado')}")

        return item
