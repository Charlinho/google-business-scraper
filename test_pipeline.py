#!/usr/bin/env python3
"""
Script de teste para verificar o funcionamento do pipeline
"""

import os
import sys
from openpyxl import Workbook

# Adicionar o diretório do projeto ao path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def test_excel_creation():
    """Teste simples para verificar se conseguimos criar um arquivo Excel"""
    print("🧪 Testando criação de arquivo Excel...")

    try:
        # Criar diretório data se não existir
        os.makedirs('data', exist_ok=True)
        print("✅ Pasta 'data' criada/verificada")

        # Criar workbook de teste
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Salões de Beleza"

        # Adicionar cabeçalhos
        headers = ['Nome', 'Avaliação', 'Número de Avaliações', 'Localização', 'Categoria', 'URL']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Adicionar dados de teste
        test_data = [
            ['Salão Teste 1', '4.5', '(123)', 'Atibaia, SP', 'Salão de Beleza', 'https://example.com/1'],
            ['Salão Teste 2', '4.8', '(89)', 'Atibaia, SP', 'Salão de Beleza', 'https://example.com/2'],
            ['Salão Teste 3', '4.2', '(45)', 'Atibaia, SP', 'Salão de Beleza', 'https://example.com/3']
        ]

        for row_num, row_data in enumerate(test_data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num, value=cell_value)

        # Salvar arquivo
        filepath = 'data/saloes_beleza_atibaia_teste.xlsx'
        workbook.save(filepath)

        print(f"✅ Arquivo Excel criado com sucesso: {filepath}")
        print(f"✅ Total de itens de teste: {len(test_data)}")

        # Verificar se o arquivo foi criado
        if os.path.exists(filepath):
            file_size = os.path.getsize(filepath)
            print(f"✅ Arquivo confirmado - Tamanho: {file_size} bytes")
            return True
        else:
            print("❌ Arquivo não foi encontrado após criação")
            return False

    except Exception as e:
        print(f"❌ Erro ao criar arquivo Excel: {e}")
        return False

def test_scrapy_items():
    """Teste para verificar se os items do Scrapy funcionam"""
    print("\n🧪 Testando BusinessItem...")

    try:
        from google_business_scraper.items import BusinessItem

        # Criar item de teste
        item = BusinessItem()
        item['name'] = 'Salão Teste'
        item['rating'] = '4.5'
        item['review_count'] = '(100)'
        item['location'] = 'Atibaia, SP'
        item['category'] = 'Salão de Beleza'
        item['url'] = 'https://example.com'

        print("✅ BusinessItem criado com sucesso")
        print(f"✅ Item: {dict(item)}")
        return True

    except Exception as e:
        print(f"❌ Erro ao criar BusinessItem: {e}")
        print("💡 Verifique se o arquivo items.py tem todos os campos definidos")
        return False

def main():
    """Executar todos os testes"""
    print("🚀 Iniciando testes do pipeline...\n")

    # Teste 1: Criação de Excel
    excel_ok = test_excel_creation()

    # Teste 2: Items do Scrapy
    items_ok = test_scrapy_items()

    print("\n📊 Resultado dos Testes:")
    print(f"Excel: {'✅ OK' if excel_ok else '❌ FALHOU'}")
    print(f"Items: {'✅ OK' if items_ok else '❌ FALHOU'}")

    if excel_ok and items_ok:
        print("\n🎉 Todos os testes passaram! O pipeline deve funcionar.")
    else:
        print("\n⚠️  Alguns testes falharam. Verifique os erros acima.")

if __name__ == '__main__':
    main()
